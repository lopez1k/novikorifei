from aiogram import Router, F, Bot
from aiogram.types import Message, FSInputFile
from aiogram.filters import Command
from datetime import datetime
from utils.keyboards.inline_builder import tickets_kb
from utils.keyboards.reply_kb import main_kb
from utils.requestsbd import create_spect_db, check_admin, get_all_users, get_sheets
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup 
from aiogram.exceptions import TelegramBadRequest
from contextlib import suppress
from openpyxl import load_workbook

acom_router = Router()

class FormCreate(StatesGroup):
    name = State()
    desc = State()
    place = State()
    date = State()

@acom_router.message(Command('apanel'))
async def apanel_cmd(msg: Message):
    if await check_admin(msg.from_user.id) == 1:
        kb = await tickets_kb()
        await msg.reply(f"<i>Список тікетів станом на</i> <b>{datetime.now().strftime('%d-%m-%Y %H:%M:%S')}</b>", reply_markup=kb)


@acom_router.message(Command("zvit"))
async def cmd_zvit(msg: Message):
    result = await get_sheets()
    fn = "data/monday.xlsx"
    wb = load_workbook(fn)
    ws = wb['users']

    start_cell = 'A2'

    for index, res in enumerate(result):
        name_cell = ws.cell(row=int(start_cell[1:]) + index, column=1, value=res[0])  
        phone_cell = ws.cell(row=int(start_cell[1:]) + index, column=2, value=res[1])  
    wb.save(fn)

    await msg.reply_document(document = FSInputFile("data/monday.xlsx"))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None   
    wb.save(fn)
    wb.close()


@acom_router.message(Command('new_spect'))
async def create_spect(msg: Message, state: FSMContext):
    if await check_admin(msg.from_user.id) == 1:
        await msg.reply("Введіть назву вистави:")
        await state.set_state(FormCreate.name)



@acom_router.message(FormCreate.name)
async def create_spect_name(msg: Message, state: FSMContext):
    await state.update_data(name = msg.text)
    await state.set_state(FormCreate.desc)
    await msg.reply("Тепер введіть опис вистави.")


@acom_router.message(FormCreate.desc)
async def create_spect_desc(msg: Message, state: FSMContext):
    await state.update_data(desc = msg.text)
    await state.set_state(FormCreate.place)
    await msg.reply("Тепер введіть місце проведення вистави.")


@acom_router.message(FormCreate.place)
async def create_spect_place(msg: Message, state: FSMContext):
    await state.update_data(place = msg.text)
    await state.set_state(FormCreate.date)
    await msg.reply("Тепер введіть дату проведення вистави. У форматі: dd-mm-yyyy H:M (01-02-2024 17:00)")

@acom_router.message(FormCreate.date)
async def create_spect_place(msg: Message, bot: Bot, state: FSMContext):
    data = await state.get_data()
    name = data['name']
    desc = data['desc']
    place = data['place']
    date = msg.text
    await state.clear()
    await create_spect_db(name, desc, place, date)
    await msg.reply("Ви успішно створили нову виставу.", reply_markup = main_kb)
    users = await get_all_users()
    for user in users:
        with suppress(TelegramBadRequest):
            await bot.send_message(chat_id=user[1], text = f'<b>Ми підготували нову виставу:</b> <i>{name}</i>. <b>Хутчіше заходьте у розділ "Найближчі вистави".</b>')

